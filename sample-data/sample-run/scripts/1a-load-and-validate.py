"""
Data preparation script for triangle analysis.

Reads raw triangle data, optional prior selections, and optional expected loss rates,
validates them, and saves to standardized format.

Run from scripts/ directory: python 1a-load-and-validate.py
"""

import pandas as pd
from pathlib import Path
from typing import Optional

from modules import config
from modules.validators import (
    validate_combined_data,
    validate_prior_selections, 
    validate_expected_loss_rates
)

# Paths from modules/config.py — override here if needed:
DATA_FILE_PATH = config.RAW_DATA
OUTPUT_PATH = config.PROCESSED_DATA
EXPECTED_LOSS_RATES_FILE = None  # Optional: path to expected loss rates file


# =============================================================================
# TODO: IMPLEMENT DATA READING FUNCTIONS BELOW
# =============================================================================

def read_triangle_data(
    file_path: str,
    sheet_name: Optional[str] = None,
    **kwargs
) -> pd.DataFrame:
    """
    *** IMPLEMENT THIS FUNCTION TO READ YOUR RAW TRIANGLE DATA ***
    
    This function must read triangle data from your source and return it in the
    standardized long format required by the validation and downstream scripts.
    
    Required Output Format:
    -----------------------
    Returns a DataFrame with these EXACT columns:
        - period (ordered categorical): Accident/policy periods (e.g., "2020", "2021 Q1")
        - age (ordered categorical): Development ages (e.g., "12", "24", "36")
        - value (float): Numeric values from the triangle
        - measure (categorical): Data type - one of:
            * "Incurred Loss"
            * "Paid Loss"
            * "Reported Count"
            * "Closed Count"
            * "Exposure"
        - unit_type (categorical): Either "Count" or "Dollars"
        - source (categorical): Source identifier for auditing
        - details (object/str): Additional context (optional)
    
    Count Triangle Interpretation:
    ------------------------------
    **Default Assumption:** Count triangles should be assumed to represent 
    "Reported Count" unless there is strong evidence otherwise.
    
    Common patterns that indicate Reported Count:
    - Column headers like "Ct", "Count", "Claims", "# Claims"
    - Generic count data without "Closed" qualifier
    - Counts that increase or stabilize over development periods
    
    Only classify as "Closed Count" if you see explicit evidence:
    - Column headers containing "Closed", "Settled", "Finalized"
    - Source documentation clearly indicating closed/settled claims
    - Counts that plateau at mature ages (closed cannot exceed reported)
    
    Special Handling for Exposure:
    ------------------------------
    **Recommended:** Provide Exposure in simple 2-column format (period, value)
    - Set age=None for all Exposure rows
    - One row per period
    - Downstream scripts only use one value per period anyway
    
    If your raw data has Exposure as a triangle:
    - Extract the LATEST DIAGONAL yourself (most mature age for each period)
    - Do NOT pass the full triangle
    
    Example (recommended):
        period="2020", age=None, value=1000, measure="Exposure", unit_type="Count"
        period="2021", age=None, value=1200, measure="Exposure", unit_type="Count"
    
    Example Usage:
    --------------
    incurred = read_triangle_data(
        file_path="my_data.xlsx",
        sheet_name="Incurred",
        measure="Incurred Loss",
        unit_type="Dollars"
    )
    
    Args:
        file_path: Path to your data file
        sheet_name: Sheet name if Excel file
        **kwargs: Additional parameters you need (measure, unit_type, etc.)
    
    Returns:
        DataFrame in the format described above
    """
    import openpyxl
    sheet_name = kwargs.get('sheet_name', sheet_name)
    measure = kwargs.get('measure', 'Paid Loss')
    unit_type = kwargs.get('unit_type', 'Dollars')
    source = kwargs.get('source', 'Triangle Examples 1.xlsx')

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Row 0: "Age of Evaluation" label row (Paid 1, Inc 1) or "Accident Year" + ages (Ct 1)
    # Detect format: if first cell of row 0 is "Accident Year", ages are in row 0; data starts row 1
    # Otherwise (Paid/Inc format): ages in row 1 col 1+; data starts row 2
    if str(rows[0][0]).strip() == 'Accident Year':
        age_row = rows[0]
        data_start = 1
    else:
        age_row = rows[1]
        data_start = 2

    ages = [str(int(v)) for v in age_row[1:] if v is not None]
    records = []
    for row in rows[data_start:]:
        ay = row[0]
        if ay is None:
            continue
        period = str(int(ay))
        for i, age in enumerate(ages):
            v = row[i + 1]
            if v is not None:
                records.append({
                    'period': period,
                    'age': age,
                    'value': float(v),
                    'measure': measure,
                    'unit_type': unit_type,
                    'source': source,
                    'details': ''
                })

    df = pd.DataFrame(records)
    periods_sorted = sorted(df['period'].unique(), key=lambda x: int(x))
    ages_sorted = sorted(df['age'].unique(), key=lambda x: int(x))
    df['period'] = pd.Categorical(df['period'], categories=periods_sorted, ordered=True)
    df['age'] = pd.Categorical(df['age'], categories=ages_sorted, ordered=True)
    df['measure'] = df['measure'].astype('category')
    df['unit_type'] = df['unit_type'].astype('category')
    df['source'] = df['source'].astype('category')
    df = df.sort_values(['period', 'age']).reset_index(drop=True)
    return df


def read_and_process_triangles():
    """
    *** IMPLEMENT THIS FUNCTION TO PROCESS YOUR TRIANGLES ***
    
    This function should:
    1. Call read_triangle_data() for each triangle type you have
    2. Combine them into a single DataFrame
    3. Ensure proper categorical ordering (CRITICAL: age must be ordered categorical)
    4. Pass validation
    5. Save to CSV

    Example Implementation:
    -----------------------
    # Read each triangle type
    incurred = read_triangle_data(
        file_path=DATA_FILE_PATH + "my_data.xlsx",
        sheet_name="Incurred",
        measure="Incurred Loss",
        unit_type="Dollars"
    )

    paid = read_triangle_data(...)

    # For Exposure: create with age=None, but define age categories from other triangles
    exposure_data = []
    for period in incurred['period'].cat.categories:
        exposure_data.append({
            'period': period,
            'age': None,  # Exposure doesn't develop
            'value': get_exposure_for_period(period),
            'measure': 'Exposure',
            'unit_type': 'Count',
            'source': 'your_source',
            'details': ''
        })
    exposure = pd.DataFrame(exposure_data)

    # Combine
    all_data = pd.concat([incurred, paid, exposure, ...], ignore_index=True)

    # CRITICAL: Re-apply categorical ordering after concat
    # The age column MUST be an ordered categorical even though Exposure rows have None values
    all_data['period'] = pd.Categorical(
        all_data['period'],
        categories=incurred['period'].cat.categories,
        ordered=True
    )
    all_data['age'] = pd.Categorical(
        all_data['age'],  # This will have None values for Exposure rows
        categories=incurred['age'].cat.categories,  # Use age categories from a non-Exposure triangle
        ordered=True
    )
    all_data['measure'] = all_data['measure'].astype('category')
    all_data['unit_type'] = all_data['unit_type'].astype('category')
    all_data['source'] = all_data['source'].astype('category')

    # Validate using the combined validator that handles both triangles and exposure
    validate_combined_data(all_data)

    # Save — row order matters. Downstream scripts reconstruct age/period sort order from
    # first occurrence in this file, so write rows in chronological period / ascending age order.
    all_data.to_csv(OUTPUT_PATH + "1_triangles.csv", index=False)
    """
    excel_path = DATA_FILE_PATH + "Triangle Examples 1.xlsx"

    paid = read_triangle_data(excel_path, sheet_name='Paid 1',
        measure='Paid Loss', unit_type='Dollars', source='Triangle Examples 1.xlsx')
    incurred = read_triangle_data(excel_path, sheet_name='Inc 1',
        measure='Incurred Loss', unit_type='Dollars', source='Triangle Examples 1.xlsx')
    counts = read_triangle_data(excel_path, sheet_name='Ct 1',
        measure='Reported Count', unit_type='Count', source='Triangle Examples 1.xlsx')

    # Exposure: read from Exposure sheet, one row per period
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Exposure']
    exp_rows = list(ws.iter_rows(min_row=2, values_only=True))
    exp_data = []
    for r in exp_rows:
        if r[0] is not None and r[1] is not None:
            exp_data.append({
                'period': str(int(r[0])),
                'age': None,
                'value': float(r[1]),
                'measure': 'Exposure',
                'unit_type': 'Count',
                'source': 'Triangle Examples 1.xlsx',
                'details': ''
            })
    exposure = pd.DataFrame(exp_data)

    all_data = pd.concat([paid, incurred, counts, exposure], ignore_index=True)

    periods_sorted = list(paid['period'].cat.categories)
    ages_sorted = list(paid['age'].cat.categories)

    all_data['period'] = pd.Categorical(all_data['period'], categories=periods_sorted, ordered=True)
    all_data['age'] = pd.Categorical(all_data['age'], categories=ages_sorted, ordered=True)
    all_data['measure'] = all_data['measure'].astype('category')
    all_data['unit_type'] = all_data['unit_type'].astype('category')
    all_data['source'] = all_data['source'].astype('category')

    all_data = all_data.sort_values(['measure', 'period', 'age']).reset_index(drop=True)

    validate_combined_data(all_data)
    all_data.to_csv(OUTPUT_PATH + "1_triangles.csv", index=False)
    print(f"  Saved {len(all_data)} rows to {OUTPUT_PATH}1_triangles.csv")


def read_and_process_prior_selections(triangle_data: pd.DataFrame) -> Optional[pd.DataFrame]:
    """Read and validate prior selections if they exist."""
    prior_file = OUTPUT_PATH + "../prior-selections.csv"
    
    if not Path(prior_file).exists():
        print("  No prior selections file found (optional)")
        return None
    
    print(f"  Found prior selections file: {prior_file}")
    df_prior = pd.read_csv(prior_file)
    
    # Ensure required columns
    required = ['measure', 'interval', 'selection']
    if not all(col in df_prior.columns for col in required):
        raise ValueError(f"Prior selections must have columns: {', '.join(required)}")
    
    if 'reasoning' not in df_prior.columns:
        df_prior['reasoning'] = ''
    
    if df_prior.empty:
        print("  Prior selections file is empty")
        return None
    
    validate_prior_selections(df_prior, triangle_data)
    
    # Ensure consistent types
    df_prior = df_prior[['measure', 'interval', 'selection', 'reasoning']].copy()
    df_prior['measure'] = df_prior['measure'].astype(str)
    df_prior['interval'] = df_prior['interval'].astype(str)
    df_prior['selection'] = df_prior['selection'].astype(float)
    df_prior['reasoning'] = df_prior['reasoning'].fillna('').astype(str)
    
    print(f"  Validated {len(df_prior)} prior selections")
    return df_prior



def read_and_process_expected_loss_rates(triangle_data: pd.DataFrame, file_path: Optional[str] = None) -> Optional[pd.DataFrame]:
    """Read and validate expected loss rates if they exist."""
    if file_path is None:
        if EXPECTED_LOSS_RATES_FILE is None:
            print("  Expected loss rates not configured")
            return None
        file_path = DATA_FILE_PATH + EXPECTED_LOSS_RATES_FILE
    
    if not Path(file_path).exists():
        print("  No expected loss rates file found (optional)")
        return None
    
    print(f"  Found expected loss rates file: {file_path}")
    
    # Read file
    if file_path.lower().endswith('.csv'):
        df = pd.read_csv(file_path)
    elif file_path.lower().endswith(('.xlsx', '.xls')):
        df = pd.read_excel(file_path, engine='openpyxl', engine_kwargs={'data_only': True})
    else:
        raise ValueError(f"Unsupported file format: {file_path}")
    
    # Clean data
    df = df.dropna(subset=['period', 'expected_loss_rate', 'expected_freq'], how='all')
    df = df.dropna(subset=['period'])
    df['period'] = df['period'].astype(str).str.strip()
    df = df.dropna(subset=['expected_loss_rate', 'expected_freq'], how='all')

    df['expected_loss_rate'] = pd.to_numeric(df['expected_loss_rate'], errors='coerce')
    df['expected_freq'] = pd.to_numeric(df['expected_freq'], errors='coerce')

    df.to_csv(OUTPUT_PATH + "1_expected_loss_rates.csv", index=False)
    print(f"  Saved {len(df)} expected loss rate records")
    return df

if __name__ == "__main__":
    """Run the data preparation process."""
    print("="*70)
    print("TRIANGLE DATA PREPARATION SCRIPT")
    print("="*70)
    
    try:
        # Process triangle data
        print("\n[1/3] Processing triangle data...")
        read_and_process_triangles()
        print(f"✓ Triangle data complete: {OUTPUT_PATH}1_triangles.csv")

        # Load triangles (already validated in read_and_process_triangles)
        df_triangles = pd.read_csv(OUTPUT_PATH + "1_triangles.csv")
        print(f"  ✓ Loaded {len(df_triangles)} rows")
        
        # Process prior selections (optional)
        print("\n[2/3] Processing prior selections (if available)...")
        df_prior = read_and_process_prior_selections(df_triangles)
        if df_prior is not None:
            output_file = OUTPUT_PATH + "../prior-selections.csv"
            df_prior.to_csv(output_file, index=False)
            print(f"✓ Prior selections validated and saved to: {output_file}")
        
        # Process expected loss rates (optional)
        print("\n[3/3] Processing expected loss rates (if available)...")
        df_expected = read_and_process_expected_loss_rates(df_triangles)
        if df_expected is not None:
            validate_expected_loss_rates(df_expected, df_triangles)
            print(f"✓ Expected loss rates validated and saved to: {OUTPUT_PATH}1_expected_loss_rates.csv")
        
        print("\n" + "="*70)
        print("✓ DATA PREPARATION COMPLETE!")
        print("="*70)
        
    except NotImplementedError as e:
        print("\n" + str(e))
        print("\nNEXT STEPS:")
        print("1. Implement read_triangle_data() to read your raw data source")
        print("2. Implement read_and_process_triangles() to combine triangles")
        print("3. Ensure output passes validate_combined_data() checks")
        print("4. Run this script again")
        print("\nSee function docstrings for detailed requirements.")
        print("="*70)
        exit(1)
    except Exception as e:
        print(f"\n✗ ERROR: {type(e).__name__}: {e}")
        print("="*70)
        exit(1)
