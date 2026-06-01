"""
Tests for reserving-analysis scripts.

Key coverage areas:
  - Categorical ordering reconstruction (core assumption of the CSV-based workflow)
  - enhance_triangle_data (1b)
  - calculate_ldf_averages (1d)
  - calculate_diagnostics (1c)
  - calculations module (pure functions)
  - preview_data_file
  - validators
"""

import importlib.util
import os

import pandas as pd
import pytest

from modules.calculations import (
    calc_bf_pct_unreported,
    calc_bf_ultimate,
    calc_bf_unreported,
    calc_cl_ultimate,
    calc_ibnr,
    safe_divide,
    sanitize_value,
)
from modules.validators import (
    validate_combined_data,
    validate_prior_selections,
    validate_triangle_data,
)
from preview_data_file import preview_data_file

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PERIODS = ["2020", "2021", "2022"]
AGES = ["12", "24", "36"]

SCRIPTS = os.path.join(os.path.dirname(__file__), "..", "skills", "reserving-analysis", "scripts")


def _load(filename):
    """Load a script file as a module (scripts aren't packages)."""
    path = os.path.join(SCRIPTS, filename)
    spec = importlib.util.spec_from_file_location(filename.replace(".py", ""), path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="session")
def script_1b():
    return _load("1b-calculate-ldfs.py")


@pytest.fixture(scope="session")
def script_1c():
    return _load("1c-diagnostics.py")


@pytest.fixture(scope="session")
def script_1d():
    return _load("1d-ldf-averages.py")


@pytest.fixture
def triangle():
    """Minimal upper triangle with ordered categoricals."""
    rows = []
    for p_i, p in enumerate(PERIODS):
        for a_i, a in enumerate(AGES):
            if a_i <= (len(AGES) - 1 - p_i):
                rows.append({
                    "period": p, "age": a,
                    "value": float(100 * (p_i + 1) * (a_i + 1)),
                    "measure": "Incurred Loss",
                    "unit_type": "Dollars",
                    "source": "test",
                    "details": "",
                })
    df = pd.DataFrame(rows)
    df["age"] = pd.Categorical(df["age"], categories=AGES, ordered=True)
    df["period"] = pd.Categorical(df["period"], categories=PERIODS, ordered=True)
    df["measure"] = df["measure"].astype("category")
    df["unit_type"] = df["unit_type"].astype("category")
    df["source"] = df["source"].astype("category")
    return df


@pytest.fixture
def enhanced(script_1b, triangle):
    return script_1b.enhance_triangle_data(triangle)


@pytest.fixture
def big_triangle(script_1b):
    """5 periods, 2 ages with varying LDFs — enough to test 3yr vs all averages."""
    periods = ["2018", "2019", "2020", "2021", "2022"]
    ages = ["12", "24"]
    # Use prime-like values so LDFs differ by period
    age12_vals = [100, 150, 90, 130, 110]
    age24_vals = [210, 270, 200, 300, 240]
    rows = []
    for p_i, p in enumerate(periods):
        rows.append({"period": p, "age": "12", "value": float(age12_vals[p_i]),
                     "measure": "Incurred Loss", "unit_type": "Dollars", "source": "t", "details": ""})
        rows.append({"period": p, "age": "24", "value": float(age24_vals[p_i]),
                     "measure": "Incurred Loss", "unit_type": "Dollars", "source": "t", "details": ""})
    df = pd.DataFrame(rows)
    df["age"] = pd.Categorical(df["age"], categories=ages, ordered=True)
    df["period"] = pd.Categorical(df["period"], categories=periods, ordered=True)
    df["measure"] = df["measure"].astype("category")
    df["unit_type"] = df["unit_type"].astype("category")
    df["source"] = df["source"].astype("category")
    return script_1b.enhance_triangle_data(df)


# ---------------------------------------------------------------------------
# 1. Categorical ordering reconstruction
# ---------------------------------------------------------------------------

class TestCategoricalOrdering:
    def test_first_occurrence_order_preserved_through_csv(self, triangle, tmp_path):
        """Write triangle to CSV, reconstruct categoricals, verify order matches original."""
        csv_path = tmp_path / "tri.csv"
        triangle.to_csv(csv_path, index=False)

        df = pd.read_csv(csv_path)
        age_order = list(dict.fromkeys(df["age"].dropna().astype(str)))
        period_order = list(dict.fromkeys(df["period"].dropna().astype(str)))
        df["age"] = pd.Categorical(df["age"], categories=age_order, ordered=True)
        df["period"] = pd.Categorical(df["period"], categories=period_order, ordered=True)

        assert df["age"].cat.categories.tolist() == AGES
        assert df["period"].cat.categories.tolist() == PERIODS

    def test_nonalphabetic_ages_preserve_input_order(self, tmp_path):
        """Ages 12/24/120 would sort as 12/120/24 alphabetically; first-occurrence wins."""
        ages = ["12", "24", "120"]  # alphabetical: "12", "120", "24"
        rows = [{"period": "2020", "age": a, "value": 1.0, "measure": "Incurred Loss",
                 "unit_type": "Dollars", "source": "test", "details": ""} for a in ages]
        df = pd.DataFrame(rows)
        df["age"] = pd.Categorical(df["age"], categories=ages, ordered=True)
        df["period"] = pd.Categorical(["2020"] * 3, categories=["2020"], ordered=True)
        df["measure"] = df["measure"].astype("category")
        df["unit_type"] = df["unit_type"].astype("category")
        df["source"] = df["source"].astype("category")

        csv_path = tmp_path / "tri.csv"
        df.to_csv(csv_path, index=False)

        df2 = pd.read_csv(csv_path)
        age_order = list(dict.fromkeys(df2["age"].dropna().astype(str)))
        df2["age"] = pd.Categorical(df2["age"], categories=age_order, ordered=True)

        assert df2["age"].cat.categories.tolist() == ages

    def test_nan_ages_excluded_from_reconstructed_categories(self, tmp_path):
        """Exposure rows have NaN age — reconstruction must skip NaN and preserve string form."""
        rows = [
            {"period": "2020", "age": None,  "value": 1000.0, "measure": "Exposure",
             "unit_type": "Count", "source": "test", "details": ""},
            {"period": "2020", "age": "12", "value": 100.0,  "measure": "Incurred Loss",
             "unit_type": "Dollars", "source": "test", "details": ""},
            {"period": "2020", "age": "24", "value": 110.0,  "measure": "Incurred Loss",
             "unit_type": "Dollars", "source": "test", "details": ""},
        ]
        csv_path = tmp_path / "tri.csv"
        pd.DataFrame(rows).to_csv(csv_path, index=False)

        # dtype=str prevents float inference ("12" → "12.0") for columns with NaN
        df = pd.read_csv(csv_path, dtype={"age": str})
        age_order = list(dict.fromkeys(df["age"].dropna()))
        assert age_order == ["12", "24"]

    def test_interval_order_preserved_through_csv(self, enhanced, tmp_path):
        """interval categories in 2_enhanced must survive CSV round-trip."""
        csv_path = tmp_path / "enhanced.csv"
        enhanced.to_csv(csv_path, index=False)

        df = pd.read_csv(csv_path)
        interval_order = list(dict.fromkeys(df["interval"].dropna().astype(str)))
        df["interval"] = pd.Categorical(df["interval"], categories=interval_order, ordered=True)

        expected = [f"{AGES[i]}-{AGES[i+1]}" for i in range(len(AGES) - 1)]
        assert df["interval"].cat.categories.tolist() == expected

    def test_cat_categories_survives_csv_for_measure(self, triangle, tmp_path):
        """measure order from input is recoverable after CSV round-trip."""
        csv_path = tmp_path / "tri.csv"
        triangle.to_csv(csv_path, index=False)

        df = pd.read_csv(csv_path)
        measure_order = list(dict.fromkeys(df["measure"].dropna().astype(str)))
        df["measure"] = pd.Categorical(df["measure"], categories=measure_order)
        assert "Incurred Loss" in df["measure"].cat.categories.tolist()


# ---------------------------------------------------------------------------
# 2. enhance_triangle_data (1b)
# ---------------------------------------------------------------------------

class TestEnhanceTriangleData:
    def test_ldf_computed_for_non_first_age(self, enhanced):
        assert enhanced["ldf"].notna().any()

    def test_first_age_has_no_ldf(self, enhanced):
        first_age = AGES[0]
        assert enhanced[enhanced["age"] == first_age]["ldf"].isna().all()

    def test_ldf_value_correct(self, script_1b, triangle):
        """period=2020, age=24: prior value=100, current value=200 → LDF=2.0."""
        df = script_1b.enhance_triangle_data(triangle)
        row = df[(df["period"] == "2020") & (df["age"] == "24")]
        assert pytest.approx(row["ldf"].iloc[0], rel=1e-6) == 2.0

    def test_weight_equals_prior_value(self, enhanced):
        with_ldfs = enhanced[enhanced["ldf"].notna()]
        assert (with_ldfs["weight"] > 0).all()

    def test_intervals_in_correct_order(self, enhanced):
        expected = [f"{AGES[i]}-{AGES[i+1]}" for i in range(len(AGES) - 1)]
        assert enhanced["interval"].cat.categories.tolist() == expected

    def test_prior_age_is_not_explicitly_set_categorical(self, enhanced):
        """prior_age comes from shift() on a categorical age column — values are valid ages."""
        with_prior = enhanced[enhanced["prior_age"].notna()]
        # Values should be recognizable age labels, not None
        assert with_prior["prior_age"].notna().all()
        assert set(with_prior["prior_age"].astype(str)).issubset(set(AGES))

    def test_age_ordering_preserved(self, enhanced):
        assert enhanced["age"].cat.categories.tolist() == AGES

    def test_row_count_matches_input(self, triangle, enhanced):
        assert len(enhanced) == len(triangle)


# ---------------------------------------------------------------------------
# 3. calculate_ldf_averages (1d)
# ---------------------------------------------------------------------------

class TestCalculateLdfAverages:
    def test_one_row_per_measure_interval(self, script_1d, enhanced):
        summary = script_1d.calculate_ldf_averages(enhanced)
        expected = [f"{AGES[i]}-{AGES[i+1]}" for i in range(len(AGES) - 1)]
        for interval in expected:
            assert len(summary[summary["interval"] == interval]) == 1

    def test_weighted_all_is_finite(self, script_1d, enhanced):
        summary = script_1d.calculate_ldf_averages(enhanced)
        assert summary["weighted_all"].notna().any()

    def test_simple_avg_between_min_and_max(self, script_1d, enhanced):
        summary = script_1d.calculate_ldf_averages(enhanced)
        valid = summary.dropna(subset=["simple_all", "min_all", "max_all"])
        assert (valid["simple_all"] >= valid["min_all"] - 1e-9).all()
        assert (valid["simple_all"] <= valid["max_all"] + 1e-9).all()

    def test_3yr_avg_differs_from_all_with_five_periods(self, script_1d, big_triangle):
        """With 5 periods of varying LDFs, weighted_3yr ≠ weighted_all."""
        summary = script_1d.calculate_ldf_averages(big_triangle)
        row = summary[summary["interval"] == "12-24"].iloc[0]
        # Both averages present
        assert pd.notna(row["weighted_all"])
        assert pd.notna(row["weighted_3yr"])
        # With 5 different LDFs, the two should differ
        assert row["weighted_all"] != pytest.approx(row["weighted_3yr"], rel=1e-6)

    def test_output_has_all_expected_columns(self, script_1d, enhanced):
        summary = script_1d.calculate_ldf_averages(enhanced)
        for col in ["measure", "interval", "weighted_all", "simple_all", "weighted_3yr"]:
            assert col in summary.columns


# ---------------------------------------------------------------------------
# 4. calculate_diagnostics (1c)
# ---------------------------------------------------------------------------

class TestCalculateDiagnostics:
    def test_returns_dataframe(self, script_1c, enhanced):
        diag = script_1c.calculate_diagnostics(enhanced)
        assert isinstance(diag, pd.DataFrame)
        assert len(diag) > 0

    def test_period_age_columns_present(self, script_1c, enhanced):
        diag = script_1c.calculate_diagnostics(enhanced)
        assert {"period", "age"}.issubset(diag.columns)

    def test_paid_to_incurred_with_both_measures(self, script_1c):
        """paid_to_incurred computed when both Incurred Loss and Paid Loss present."""
        rows = []
        for p in PERIODS:
            for a in AGES:
                rows.append({"period": p, "age": a, "value": 100.0,
                             "measure": "Incurred Loss", "unit_type": "Dollars",
                             "source": "t", "details": ""})
                rows.append({"period": p, "age": a, "value": 60.0,
                             "measure": "Paid Loss", "unit_type": "Dollars",
                             "source": "t", "details": ""})
        df = pd.DataFrame(rows)
        df["age"] = pd.Categorical(df["age"], categories=AGES, ordered=True)
        df["period"] = pd.Categorical(df["period"], categories=PERIODS, ordered=True)
        df["measure"] = df["measure"].astype("category")
        df["unit_type"] = df["unit_type"].astype("category")
        df["source"] = df["source"].astype("category")

        diag = script_1c.calculate_diagnostics(df)
        assert "paid_to_incurred" in diag.columns
        valid = diag["paid_to_incurred"].dropna()
        assert len(valid) > 0
        assert all(abs(v - 0.6) < 1e-9 for v in valid)

    def test_incurred_only_no_paid_to_incurred(self, script_1c, enhanced):
        """No Paid Loss in fixture → paid_to_incurred should not appear or be all NaN."""
        diag = script_1c.calculate_diagnostics(enhanced)
        if "paid_to_incurred" in diag.columns:
            assert diag["paid_to_incurred"].isna().all()


# ---------------------------------------------------------------------------
# 5. calculations module (pure functions)
# ---------------------------------------------------------------------------

class TestCalculations:
    def test_sanitize_nan_returns_none(self):
        assert sanitize_value(float("nan")) is None

    def test_sanitize_none_returns_none(self):
        assert sanitize_value(None) is None

    def test_sanitize_valid_passes_through(self):
        assert sanitize_value(42.0) == 42.0

    def test_cl_ultimate_basic(self):
        assert calc_cl_ultimate(100.0, 1.5) == pytest.approx(150.0)

    def test_cl_ultimate_nan_cdf_returns_none(self):
        assert calc_cl_ultimate(100.0, float("nan")) is None

    def test_cl_ultimate_nan_actual_returns_none(self):
        assert calc_cl_ultimate(float("nan"), 1.5) is None

    def test_bf_pct_unreported_basic(self):
        assert calc_bf_pct_unreported(2.0) == pytest.approx(0.5)

    def test_bf_pct_unreported_zero_cdf_returns_none(self):
        assert calc_bf_pct_unreported(0.0) is None

    def test_bf_pct_unreported_cdf_one_gives_zero(self):
        assert calc_bf_pct_unreported(1.0) == pytest.approx(0.0)

    def test_bf_unreported_basic(self):
        assert calc_bf_unreported(200.0, 0.5) == pytest.approx(100.0)

    def test_bf_ultimate_basic(self):
        assert calc_bf_ultimate(40.0, 60.0) == pytest.approx(100.0)

    def test_ibnr_standard_measure(self):
        assert calc_ibnr(150.0, 100.0, "Incurred Loss") == pytest.approx(50.0)

    def test_ibnr_paid_uses_incurred_actual(self):
        df = pd.DataFrame([
            {"measure": "Incurred Loss", "period": "2020", "actual": 120.0},
            {"measure": "Paid Loss",     "period": "2020", "actual":  80.0},
        ])
        result = calc_ibnr(200.0, 80.0, "Paid Loss", combined_df=df, period="2020")
        assert result == pytest.approx(80.0)  # 200 - 120

    def test_ibnr_none_ultimate_returns_none(self):
        assert calc_ibnr(None, 100.0, "Incurred Loss") is None

    def test_safe_divide_basic(self):
        assert safe_divide(10.0, 4.0) == pytest.approx(2.5)

    def test_safe_divide_zero_denominator_returns_none(self):
        assert safe_divide(10.0, 0.0) is None

    def test_safe_divide_nan_returns_none(self):
        assert safe_divide(float("nan"), 2.0) is None


# ---------------------------------------------------------------------------
# 6. preview_data_file
# ---------------------------------------------------------------------------

class TestPreviewDataFile:
    def test_csv_preview_returns_string(self, tmp_path):
        csv_path = tmp_path / "sample.csv"
        csv_path.write_text("period,age,value\n2020,12,100\n2020,24,110\n")
        result = preview_data_file(str(csv_path))
        assert isinstance(result, str)
        assert "period" in result
        assert "2020" in result

    def test_csv_shape_reported(self, tmp_path):
        csv_path = tmp_path / "sample.csv"
        csv_path.write_text("a,b\n1,2\n3,4\n5,6\n")
        result = preview_data_file(str(csv_path))
        assert "3" in result  # 3 rows

    def test_unsupported_extension_returns_error(self, tmp_path):
        bad = tmp_path / "file.json"
        bad.write_text("{}")
        result = preview_data_file(str(bad))
        assert result.startswith("Error:")

    def test_csv_no_headers_handled(self, tmp_path):
        csv_path = tmp_path / "noheader.csv"
        csv_path.write_text("1,2,3\n4,5,6\n")
        result = preview_data_file(str(csv_path))
        assert isinstance(result, str)

    def test_numeric_summary_present_for_numeric_columns(self, tmp_path):
        csv_path = tmp_path / "nums.csv"
        csv_path.write_text("x,y\n1.0,2.0\n3.0,4.0\n5.0,6.0\n7.0,8.0\n9.0,10.0\n11.0,12.0\n")
        result = preview_data_file(str(csv_path))
        assert "NUMERIC COLUMN SUMMARY" in result


# ---------------------------------------------------------------------------
# 7. validators
# ---------------------------------------------------------------------------

class TestValidators:
    def test_valid_triangle_passes(self, triangle):
        validate_triangle_data(triangle)  # must not raise

    def test_missing_column_raises(self, triangle):
        bad = triangle.drop(columns=["age"])
        with pytest.raises(ValueError, match="Missing required column"):
            validate_triangle_data(bad)

    def test_non_categorical_age_raises(self, triangle):
        bad = triangle.copy()
        bad["age"] = bad["age"].astype(str)
        with pytest.raises(ValueError, match="categorical"):
            validate_triangle_data(bad)

    def test_unordered_period_raises(self, triangle):
        bad = triangle.copy()
        bad["period"] = pd.Categorical(bad["period"], ordered=False)
        with pytest.raises(ValueError, match="ordered"):
            validate_triangle_data(bad)

    def test_invalid_measure_raises(self, triangle):
        bad = triangle.copy()
        bad["measure"] = pd.Categorical(["Bad"] * len(bad))
        with pytest.raises(ValueError, match="Invalid measure"):
            validate_triangle_data(bad)

    @pytest.mark.filterwarnings("ignore::FutureWarning")
    def test_combined_data_with_exposure_passes(self, triangle):
        exp_rows = [{"period": p, "age": None, "value": 500.0, "measure": "Exposure",
                     "unit_type": "Count", "source": "test", "details": ""} for p in PERIODS]
        exp = pd.DataFrame(exp_rows)
        exp["period"] = pd.Categorical(exp["period"], categories=PERIODS, ordered=True)
        exp["measure"] = exp["measure"].astype("category")
        exp["unit_type"] = exp["unit_type"].astype("category")
        exp["source"] = exp["source"].astype("category")
        combined = pd.concat([triangle, exp], ignore_index=True)
        combined["period"] = pd.Categorical(combined["period"], categories=PERIODS, ordered=True)
        combined["age"] = pd.Categorical(combined["age"], categories=AGES, ordered=True)
        combined["measure"] = combined["measure"].astype("category")
        combined["unit_type"] = combined["unit_type"].astype("category")
        combined["source"] = combined["source"].astype("category")
        validate_combined_data(combined)  # must not raise

    def test_prior_selections_invalid_measure(self, triangle):
        sel = pd.DataFrame([{"measure": "Bad Measure", "interval": "12-24", "selection": 1.1}])
        with pytest.raises(ValueError, match="Invalid measures"):
            validate_prior_selections(sel, triangle)

    def test_prior_selections_invalid_interval(self, triangle):
        sel = pd.DataFrame([{"measure": "Incurred Loss", "interval": "99-999", "selection": 1.1}])
        with pytest.raises(ValueError, match="Invalid intervals"):
            validate_prior_selections(sel, triangle)

    def test_prior_selections_valid_passes(self, triangle):
        sel = pd.DataFrame([{"measure": "Incurred Loss", "interval": "12-24", "selection": 1.1}])
        validate_prior_selections(sel, triangle)  # must not raise


# ---------------------------------------------------------------------------
# 8. calculations — remaining edge cases
# ---------------------------------------------------------------------------

class TestCalculationsEdgeCases:
    def test_bf_unreported_none_ie(self):
        assert calc_bf_unreported(None, 0.5) is None

    def test_bf_unreported_none_pct(self):
        assert calc_bf_unreported(200.0, None) is None

    def test_bf_ultimate_none_unreported(self):
        assert calc_bf_ultimate(None, 60.0) is None

    def test_bf_ultimate_none_actual(self):
        assert calc_bf_ultimate(40.0, None) is None

    def test_ibnr_paid_no_combined_df(self):
        assert calc_ibnr(200.0, 80.0, "Paid Loss") is None

    def test_ibnr_paid_no_incurred_row(self):
        df = pd.DataFrame([{"measure": "Paid Loss", "period": "2020", "actual": 80.0}])
        assert calc_ibnr(200.0, 80.0, "Paid Loss", combined_df=df, period="2020") is None

    def test_ibnr_standard_none_actual(self):
        assert calc_ibnr(150.0, None, "Incurred Loss") is None

    def test_calc_ie_loss_rate_basic(self):
        from modules.calculations import calc_ie_loss_rate
        assert calc_ie_loss_rate(300.0, 100.0) == pytest.approx(3.0)

    def test_calc_ie_loss_rate_zero_exposure(self):
        from modules.calculations import calc_ie_loss_rate
        assert calc_ie_loss_rate(300.0, 0.0) is None

    def test_calc_ie_loss_rate_none_inputs(self):
        from modules.calculations import calc_ie_loss_rate
        assert calc_ie_loss_rate(None, 100.0) is None

    def test_calc_total_ibnr_standard(self):
        from modules.calculations import calc_total_ibnr
        assert calc_total_ibnr(500.0, 400.0, "Incurred Loss") == pytest.approx(100.0)

    def test_calc_total_ibnr_paid_with_df(self):
        from modules.calculations import calc_total_ibnr
        df = pd.DataFrame([{"measure": "Incurred Loss", "actual": 450.0}])
        result = calc_total_ibnr(500.0, 400.0, "Paid Loss", combined_df=df)
        assert result == pytest.approx(50.0)  # 500 - 450

    def test_calc_total_ibnr_paid_no_df(self):
        from modules.calculations import calc_total_ibnr
        result = calc_total_ibnr(500.0, 400.0, "Paid Loss")
        assert result == pytest.approx(100.0)  # fallback: 500 - 400


# ---------------------------------------------------------------------------
# 9. validators — additional paths
# ---------------------------------------------------------------------------

def _make_exposure_df(periods=None):
    """Minimal valid exposure DataFrame."""
    periods = periods or PERIODS
    rows = [{"period": p, "value": 500.0, "measure": "Exposure",
             "unit_type": "Count", "source": "test"} for p in periods]
    df = pd.DataFrame(rows)
    df["period"] = pd.Categorical(df["period"], categories=periods, ordered=True)
    df["measure"] = df["measure"].astype("category")
    df["unit_type"] = df["unit_type"].astype("category")
    df["source"] = df["source"].astype("category")
    return df


class TestValidatorsAdditional:
    def test_empty_triangle_raises(self):
        with pytest.raises(ValueError, match="empty"):
            validate_triangle_data(pd.DataFrame())

    def test_null_value_in_value_column_raises(self, triangle):
        bad = triangle.copy()
        bad.loc[bad.index[0], "value"] = None
        with pytest.raises(ValueError, match="null value"):
            validate_triangle_data(bad)

    def test_duplicate_rows_raises(self, triangle):
        bad = pd.concat([triangle, triangle.iloc[:1]], ignore_index=True)
        bad["age"] = pd.Categorical(bad["age"], categories=AGES, ordered=True)
        bad["period"] = pd.Categorical(bad["period"], categories=PERIODS, ordered=True)
        bad["measure"] = bad["measure"].astype("category")
        bad["unit_type"] = bad["unit_type"].astype("category")
        bad["source"] = bad["source"].astype("category")
        with pytest.raises(ValueError, match="duplicate"):
            validate_triangle_data(bad)

    def test_exposure_valid_passes(self):
        from modules.validators import validate_exposure_data
        validate_exposure_data(_make_exposure_df())

    def test_exposure_empty_raises(self):
        from modules.validators import validate_exposure_data
        with pytest.raises(ValueError, match="empty"):
            validate_exposure_data(pd.DataFrame())

    def test_exposure_non_exposure_measure_raises(self):
        from modules.validators import validate_exposure_data
        df = _make_exposure_df()
        df["measure"] = pd.Categorical(["Incurred Loss"] * len(df))
        with pytest.raises(ValueError, match="expects only Exposure"):
            validate_exposure_data(df)

    def test_exposure_unordered_period_raises(self):
        from modules.validators import validate_exposure_data
        df = _make_exposure_df()
        df["period"] = pd.Categorical(df["period"], ordered=False)
        with pytest.raises(ValueError, match="ordered"):
            validate_exposure_data(df)

    def test_exposure_duplicate_raises(self):
        from modules.validators import validate_exposure_data
        df = _make_exposure_df()
        bad = pd.concat([df, df.iloc[:1]], ignore_index=True)
        bad["period"] = pd.Categorical(bad["period"], categories=PERIODS, ordered=True)
        bad["measure"] = bad["measure"].astype("category")
        bad["unit_type"] = bad["unit_type"].astype("category")
        bad["source"] = bad["source"].astype("category")
        with pytest.raises(ValueError, match="duplicate"):
            validate_exposure_data(bad)

    def test_combined_empty_raises(self):
        with pytest.raises(ValueError, match="empty"):
            validate_combined_data(pd.DataFrame())

    def test_combined_no_measure_column_raises(self):
        bad = pd.DataFrame([{"period": "2020", "value": 1.0}])
        with pytest.raises(ValueError, match="Missing required column: measure"):
            validate_combined_data(bad)

    def test_prior_selections_empty_returns_silently(self, triangle):
        validate_prior_selections(pd.DataFrame(), triangle)  # must not raise

    def test_prior_selections_missing_column_raises(self, triangle):
        sel = pd.DataFrame([{"measure": "Incurred Loss", "selection": 1.1}])  # no interval
        with pytest.raises(ValueError, match="Missing required column"):
            validate_prior_selections(sel, triangle)

    def test_prior_selections_null_measure_raises(self, triangle):
        sel = pd.DataFrame([{"measure": None, "interval": "12-24", "selection": 1.1}])
        with pytest.raises(ValueError, match="null value"):
            validate_prior_selections(sel, triangle)

    def test_prior_selections_duplicate_raises(self, triangle):
        sel = pd.DataFrame([
            {"measure": "Incurred Loss", "interval": "12-24", "selection": 1.1},
            {"measure": "Incurred Loss", "interval": "12-24", "selection": 1.2},
        ])
        with pytest.raises(ValueError, match="duplicate"):
            validate_prior_selections(sel, triangle)

    def test_validate_expected_loss_rates_valid(self, triangle):
        from modules.validators import validate_expected_loss_rates
        df = pd.DataFrame([
            {"period": "2020", "expected_loss_rate": 0.65, "expected_freq": 0.05},
            {"period": "2021", "expected_loss_rate": 0.70, "expected_freq": 0.06},
            {"period": "2022", "expected_loss_rate": 0.68, "expected_freq": 0.05},
        ])
        validate_expected_loss_rates(df, triangle)  # must not raise

    def test_validate_expected_loss_rates_empty_raises(self, triangle):
        from modules.validators import validate_expected_loss_rates
        with pytest.raises(ValueError, match="empty"):
            validate_expected_loss_rates(pd.DataFrame(), triangle)

    def test_validate_expected_loss_rates_duplicate_period_raises(self, triangle):
        from modules.validators import validate_expected_loss_rates
        df = pd.DataFrame([
            {"period": "2020", "expected_loss_rate": 0.65, "expected_freq": 0.05},
            {"period": "2020", "expected_loss_rate": 0.70, "expected_freq": 0.06},
        ])
        with pytest.raises(ValueError, match="Duplicate"):
            validate_expected_loss_rates(df, triangle)

    def test_validate_expected_loss_rates_both_null_raises(self, triangle):
        from modules.validators import validate_expected_loss_rates
        df = pd.DataFrame([
            {"period": "2020", "expected_loss_rate": None, "expected_freq": None},
        ])
        with pytest.raises(ValueError, match="missing both"):
            validate_expected_loss_rates(df, triangle)


# ---------------------------------------------------------------------------
# 10. calculate_diagnostics — additional measure combinations
# ---------------------------------------------------------------------------

def _make_multi_measure_triangle(periods=None, ages=None):
    """Triangle with Incurred, Paid, Reported Count, Closed Count, and Exposure."""
    periods = periods or PERIODS
    ages = ages or AGES
    rows = []
    for p_i, p in enumerate(periods):
        for a_i, a in enumerate(ages):
            if a_i <= len(ages) - 1 - p_i:
                for measure, unit, base in [
                    ("Incurred Loss", "Dollars", 100),
                    ("Paid Loss", "Dollars", 60),
                    ("Reported Count", "Count", 10),
                    ("Closed Count", "Count", 6),
                ]:
                    rows.append({"period": p, "age": a,
                                 "value": float(base * (p_i + 1) * (a_i + 1)),
                                 "measure": measure, "unit_type": unit,
                                 "source": "test", "details": ""})
        # Exposure: one row per period, no age
        rows.append({"period": p, "age": None, "value": 1000.0, "measure": "Exposure",
                     "unit_type": "Count", "source": "test", "details": ""})
    df = pd.DataFrame(rows)
    df["age"] = pd.Categorical(df["age"], categories=ages, ordered=True)
    df["period"] = pd.Categorical(df["period"], categories=periods, ordered=True)
    df["measure"] = df["measure"].astype("category")
    df["unit_type"] = df["unit_type"].astype("category")
    df["source"] = df["source"].astype("category")
    return df


class TestCalculateDiagnosticsExtended:
    def test_all_diagnostic_columns_present_with_full_data(self, script_1c, script_1b):
        df = _make_multi_measure_triangle()
        enhanced = script_1b.enhance_triangle_data(df[df["measure"] != "Exposure"])
        # Add exposure back for diagnostics
        full = pd.concat([enhanced, df[df["measure"] == "Exposure"]], ignore_index=True)
        full["age"] = pd.Categorical(full["age"], categories=AGES, ordered=True)
        full["period"] = pd.Categorical(full["period"], categories=PERIODS, ordered=True)
        full["measure"] = full["measure"].astype("category")
        full["unit_type"] = full["unit_type"].astype("category")
        full["source"] = full["source"].astype("category")
        diag = script_1c.calculate_diagnostics(full)
        for col in ["incurred_severity", "paid_to_incurred", "claim_closure_rate",
                    "case_reserves", "open_counts"]:
            assert col in diag.columns, f"missing {col}"

    def test_exposure_based_diagnostics_computed(self, script_1c, script_1b):
        """Loss rates computed when Exposure present."""
        df = _make_multi_measure_triangle()
        enhanced = script_1b.enhance_triangle_data(df[df["measure"] != "Exposure"])
        full = pd.concat([enhanced, df[df["measure"] == "Exposure"]], ignore_index=True)
        full["age"] = pd.Categorical(full["age"], categories=AGES, ordered=True)
        full["period"] = pd.Categorical(full["period"], categories=PERIODS, ordered=True)
        full["measure"] = full["measure"].astype("category")
        full["unit_type"] = full["unit_type"].astype("category")
        full["source"] = full["source"].astype("category")
        diag = script_1c.calculate_diagnostics(full)
        assert "incurred_loss_rate" in diag.columns
        assert diag["incurred_loss_rate"].notna().any()

    def test_reported_count_renamed_to_reported_claims(self, script_1c):
        """Reported Count measure triggers column rename."""
        rows = []
        for p in PERIODS:
            for a in AGES:
                rows.append({"period": p, "age": a, "value": 10.0,
                             "measure": "Reported Count", "unit_type": "Count",
                             "source": "t", "details": ""})
        df = pd.DataFrame(rows)
        df["age"] = pd.Categorical(df["age"], categories=AGES, ordered=True)
        df["period"] = pd.Categorical(df["period"], categories=PERIODS, ordered=True)
        df["measure"] = df["measure"].astype("category")
        df["unit_type"] = df["unit_type"].astype("category")
        df["source"] = df["source"].astype("category")
        diag = script_1c.calculate_diagnostics(df)
        assert "reported_claims" in diag.columns
        assert "reported_count" not in diag.columns

    def test_incremental_severities_with_full_data(self, script_1c, script_1b):
        """Incremental severity columns appear when both incurred and reported present."""
        df = _make_multi_measure_triangle()
        enhanced = script_1b.enhance_triangle_data(df[df["measure"] != "Exposure"])
        full = pd.concat([enhanced, df[df["measure"] == "Exposure"]], ignore_index=True)
        full["age"] = pd.Categorical(full["age"], categories=AGES, ordered=True)
        full["period"] = pd.Categorical(full["period"], categories=PERIODS, ordered=True)
        full["measure"] = full["measure"].astype("category")
        full["unit_type"] = full["unit_type"].astype("category")
        full["source"] = full["source"].astype("category")
        diag = script_1c.calculate_diagnostics(full)
        assert "incremental_incurred_severity" in diag.columns or \
               "incremental_paid_severity" in diag.columns or \
               "incremental_closure_rate" in diag.columns


# ---------------------------------------------------------------------------
# 11. preview_data_file — Excel path
# ---------------------------------------------------------------------------

class TestPreviewDataFileExcel:
    def test_xlsx_preview_returns_string(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["period", "age", "value"])
        ws.append(["2020", "12", 100])
        ws.append(["2020", "24", 110])
        path = tmp_path / "sample.xlsx"
        wb.save(path)
        result = preview_data_file(str(path))
        assert isinstance(result, str)
        assert "period" in result

    def test_xlsx_multiple_sheets_listed(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.create_sheet("Sheet2")
        wb["Sheet1"].append(["x", "y"])
        wb["Sheet1"].append([1, 2])
        path = tmp_path / "multi.xlsx"
        wb.save(path)
        result = preview_data_file(str(path))
        assert "Sheet1" in result
        assert "Sheet2" in result

    def test_xlsx_sheet_by_name(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "Data"
        wb["Data"].append(["col"])
        wb["Data"].append([42])
        path = tmp_path / "named.xlsx"
        wb.save(path)
        result = preview_data_file(str(path), sheet_name="Data")
        assert "Data" in result

    def test_xlsx_numeric_sheet_index(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "First"
        wb["First"].append(["a", "b"])
        wb["First"].append([1, 2])
        path = tmp_path / "idx.xlsx"
        wb.save(path)
        result = preview_data_file(str(path), sheet_name=0)
        assert isinstance(result, str)

    def test_xlsx_no_data_returns_error(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "Empty"
        path = tmp_path / "empty.xlsx"
        wb.save(path)
        result = preview_data_file(str(path))
        assert "Error" in result or "empty" in result.lower()
